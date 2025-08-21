import os
import pandas as pd
from pathlib import Path
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import load_workbook
from src.config.postgres_config import PG_DB_URL

class StructToProcess:
    def __init__(self):
        self.engine = create_engine(PG_DB_URL, pool_size=5, max_overflow=10, pool_pre_ping=True)
        self.conn = self.engine.connect()

    def _get_file_size_mb(self, filepath):
        return os.path.getsize(filepath) / (1024 * 1024)

    def _clean_dataframe(self, df):
        df = df.dropna(how='all')   # Drop fully empty rows
        df = df.dropna(axis=1, how='all')  # Drop fully empty columns
        return df

    def _rename_duplicate_columns(self, df):
        from collections import defaultdict
        seen = defaultdict(int)
        new_columns = []
        for col in df.columns:
            if seen[col]:
                new_col = f"{col}_{seen[col]}"
            else:
                new_col = col
            seen[col] += 1
            new_columns.append(new_col)
        df.columns = new_columns
        return df

    def _insert_into_sql(self, df, table_name, mode="replace"):
        if not df.empty:
            try:
                df.to_sql(
                    table_name,
                    self.conn,
                    if_exists=mode,  # "replace" | "append"
                    index=False,
                    method="multi",  # batch insert for speed
                    chunksize=1000
                )
                print(f"✅ Inserted into table '{table_name}' ({len(df)} rows)")
            except SQLAlchemyError as e:
                print(f"❌ Error inserting into table '{table_name}': {e}")
        else:
            print(f"Skipped table '{table_name}' — cleaned DataFrame is empty")

    def _process_insert(self, df):
        if df.empty:
            print("⚠️ DataFrame is empty after cleaning, skipping insert.")
            return
        table_name = df.columns[0].lower()
        df = self._rename_duplicate_columns(df)
        df = self._clean_dataframe(df)
        self._insert_into_sql(df, table_name)


    def _load_with_pandas(self, file_path, table_name):
        encodings_to_try = ["utf-8", "cp932","shift_jis"]
        for encoding in encodings_to_try:
            try:
                df = pd.read_csv(file_path, encoding=encoding,skiprows=7)
                df = df.drop_duplicates()
                df = self._rename_duplicate_columns(df)
                df = self._clean_dataframe(df)
                self._insert_into_sql(df, table_name)
                print(f"[pandas] Loaded '{file_path}' → table '{table_name}' using encoding '{encoding}'")
                return
            except UnicodeDecodeError as e:
                print(f" Encoding '{encoding}' failed for {file_path}: {e}")
            except Exception as e:
                print(f"❌ Unexpected error for {file_path}: {e}")
                return
        print(f"❌ All encoding attempts failed for {file_path}")

    def _load_with_dask(self, file_path, table_name):
        import dask.dataframe as dd
        try:
            sample = pd.read_csv(file_path, nrows=5)
            col_dtypes = {col: "object" for col in sample.columns}
            ddf = dd.read_csv(file_path, blocksize="64MB", engine="python", dtype=col_dtypes)

            for i in range(ddf.npartitions):
                chunk = ddf.get_partition(i).compute()
                chunk = self._rename_duplicate_columns(chunk)
                mode = "append" if i > 0 else "replace"
                self._insert_into_sql(chunk, table_name, mode=mode)
            print(f"[dask] Loaded '{file_path}' → table '{table_name}'")
        except Exception as e:
            print(f"❌ Failed to load {file_path} with Dask: {e}")

    def _load_xlsx_with_pandas(self, file_path, table_name):
        try:
            wb = load_workbook(file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Handle merged cells
                merged_ranges = list(ws.merged_cells.ranges)
                for merged_range in merged_ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    value = ws.cell(row=min_row, column=min_col).value
                    if value is not None:
                        ws.unmerge_cells(str(merged_range))
                        for row in range(min_row, max_row + 1):
                            for col in range(min_col, max_col + 1):
                                ws.cell(row=row, column=col, value=value)
                df = pd.DataFrame(ws.values)
                # --- Dynamic header detection ---
                header_row = self.find_header_row(df)
                df = df.drop(range(header_row)).reset_index(drop=True)
                df.columns = df.iloc[0]
                df = df.drop(0).reset_index(drop=True)
                df = self._rename_duplicate_columns(df)
                df = df.drop_duplicates()

                full_table_name = f"{table_name}_{sheet_name}".lower()
                self._insert_into_sql(df, full_table_name)

                print(f"[xlsx] Loaded sheet '{sheet_name}' from '{file_path}' → table '{full_table_name}'")

        except Exception as e:
            print(f"❌ Failed to load Excel file {file_path}: {e}")

    def find_header_row(self, df):
        """
        Detect the header row by selecting the row with the most non-empty cells.
        Returns the index of the header row.
        """
        non_empty_counts = df.notna().sum(axis=1)  # count non-empty cells per row
        header_row = int(non_empty_counts.idxmax())  # pick the row with max non-empty cells
        return header_row
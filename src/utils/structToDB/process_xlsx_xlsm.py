import openpyxl
import json
from typing import Dict, Any, Optional, List, Tuple
from dataclasses import dataclass
from pathlib import Path
import pandas as pd
@dataclass
class FieldConfig:
    """Configuration for each field to extract"""
    name: str
    positions: List[Tuple[int, int]]  # (row, col) tuples only

def make_field_range(name: str, rows: range, col: int) -> FieldConfig:
    """Create a FieldConfig from a row range and a single column"""
    return FieldConfig(name, [(r, col) for r in rows])
def make_middle( row: int, col: int) -> FieldConfig:
    fields = []
    name=["customer","manufacturer","sevensix"]
    for i in range(3):
        rows = range(row+i*5,row+i*5+5)
        fields.extend([FieldConfig("ai_"+name[i]+"_who", [(r, col) for r in rows]),
        FieldConfig("ai_"+name[i]+"_what", [(r, col+1) for r in rows]),
        FieldConfig("ai_"+name[i]+"_when", [(r, col+5) for r in rows])])
    return fields

def make_case_fields(case_no: int, start_row: int, start_col: int = 6) -> List[FieldConfig]:
    """
    Generate FieldConfigs for a case block.
    case_no: number suffix (1,2,3,...)
    start_row: top row of the case block
    start_col: first column (default=6 for your sheet)
    """
    fields = [
        ("customer_name", (start_row, start_col)),
        ("customer_representative", (start_row, start_col+1)),
        ("occurrence_date", (start_row, start_col+2)),
        ("branch", (start_row, start_col+3)),
        ("application", (start_row+1, start_col)),
        ("manufacturer", (start_row+2, start_col)),
        ("product", (start_row+3, start_col)),
        ("competitive", (start_row+4, start_col)),
        ("amount", (start_row+5, start_col)),
        ("order_month", (start_row+6, start_col)),
        ("budget", (start_row+7, start_col)),
        ("probability", (start_row+8, start_col)),
    ]
    return [FieldConfig(f"case{case_no}_{fname}", [pos]) for fname, pos in fields]

class ExcelDataExtractor:
    """Optimized Excel data extractor (template positions only)"""
    
    def __init__(self):
        self.fields = [
            # Basic Information
            FieldConfig("reporter_name", [(2,4), (2,5), (2,6)]),
            FieldConfig("date", [(3,4), (3,5)]),
            FieldConfig("location", [(4,4), (4,5)]),
            
            # Customer Information
            make_field_range("customer_company_name", range(7, 11), 4),
            make_field_range("customer_department", range(7, 12), 5),
            make_field_range("customer_full_name", range(7, 12), 6),
            FieldConfig("customer_number", [(11,4)]),

            # Manufacturer Information
            make_field_range("manufacturer_company_name", range(7, 11), 7),
            make_field_range("manufacturer_department", range(7, 11), 8),
            make_field_range("manufacturer_full_name", range(7, 11), 9),
            FieldConfig("sevensix", [(7,10),(8,10),(9,10),(10,10)]),
            
            
        ]
        
        # Add multiple case blocks easily
        
        # define columns for middle-table
        
    
    def adjust_fields_by_dimension(self, ws):
        extra_fields = []
        print(ws.max_row)
        if ws.max_row==71:
            self.fields+=[
                # Business Information
            FieldConfig("purpose", [(12,4),(13,4),(16,4)]),
            FieldConfig("free_description", [(17,4),(18,4)]),
            FieldConfig("associated_customer_name", [(35,4)]),
            FieldConfig("competitive_information", [(36,4)]),]
            self.fields.extend(make_middle(20,5))
            self.fields += make_case_fields(case_no=1, start_row=38)
            self.fields += make_case_fields(case_no=2, start_row=48)  # example second case at row 50
            self.fields += make_case_fields(case_no=3, start_row=58)  # third case, etc.
        # check rows
        if ws.max_row == 97:
            self.fields+=[FieldConfig("annual_project_name", [(17,4),(17,5),(17,6)]),
                          FieldConfig("annual_project_schedule", [(18,4),(18,5),(18,6)]),
                          FieldConfig("annual_project_budget", [(19,4),(19,5),(19,6)]),
                          FieldConfig("free_description", [(20,4),(20,5)]),
                          FieldConfig("associated_customer_name", [(46,4)]),
                          FieldConfig("competitive_information", [(47,4)]),
                          ]
            self.fields.extend(make_middle(31,5))

            self.fields += make_case_fields(case_no=1, start_row=49)
            self.fields += make_case_fields(case_no=2, start_row=59)  # example second case at row 50
            self.fields += make_case_fields(case_no=3, start_row=69)  # third case, etc.
            self.fields += make_case_fields(case_no=4, start_row=79)
            self.fields += make_case_fields(case_no=5, start_row=89)



    # ----------- MAIN EXTRACTION ----------
    def extract_from_file(self, file_path: str) -> Dict[str, Any]:
        """Extract values only from predefined positions"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            print(f"Processing: {Path(file_path).name} ({ws.title}, {ws.max_row}x{ws.max_column})")
            self.adjust_fields_by_dimension(ws)
            extracted, stats = {}, {"found": 0, "not_found": 0}

            for field in self.fields:
                value = self._extract_field(ws, field)
                extracted[field.name] = value
                stats["found" if value else "not_found"] += 1
            
            results = {
                "data": extracted, 
                "file_info": {
                    "filename": Path(file_path).name,
                    "worksheet": ws.title,
                    "dimensions": f"{ws.max_row}x{ws.max_column}"
                },
                "extraction_stats": {
                    **stats,
                    "total_fields": len(self.fields),
                    "success_rate": f"{(stats['found']/len(self.fields))*100:.1f}%"
                }
            }
            # self._print_summary(results)
            return results
        except Exception as e:
            print(f"❌ Error: {e}")
            return {"error": str(e)}

    # ----------- FIELD EXTRACTION ----------
    def _extract_field(self, ws, field: FieldConfig) -> Optional[str]:
        """Extract value only from predefined positions"""
        values = [
            str(ws.cell(row=r, column=c).value).strip()
            for r, c in field.positions
            if ws.cell(row=r, column=c).value not in (None, "")
        ]
        if values:
            return "\n".join(values) if len(values) > 1 else values[0]
        return None

    # ----------- UTILITIES ----------
    def _print_summary(self, results: Dict[str, Any]):
        if "error" in results:
            print(f"❌ Extraction failed: {results['error']}")
            return
        stats = results["extraction_stats"]
        print(f"\n{'='*50}\nEXTRACTION SUMMARY\n{'='*50}")
        print(f"📊 Success Rate: {stats['success_rate']}")
        print(f"✅ Found: {stats['found']}")
        print(f"❌ Not found: {stats['not_found']}")
        print(f"📋 Total fields: {stats['total_fields']}")

    def save_results(self, results: Dict[str, Any], output_file: str):
        try:
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            print(f"💾 Saved: {output_file}")
        except Exception as e:
            print(f"❌ Save failed: {e}")

    
    def batch_process(self, files: List[str],level:str="1",origin:str="s3_bucket" ,output_dir: str = "outputs"):
        Path(output_dir).mkdir(exist_ok=True)
        all_results = []

        print(f"🔄 Processing {len(files)} files...")
        for i, file in enumerate(files, 1):
            print(f"\n[{i}/{len(files)}] {Path(file).name}")
            result = self.extract_from_file(file)
            
            if result and "data" in result:
                # Save JSON as before
                #self.save_results(result, str(Path(output_dir) / f"{Path(file).stem}_extracted.json"))
                row = result["data"].copy()
                row["source"] = file
                row["level"]=level
                row["origin"]=origin
                all_results.append(row)
        # Save all results to combined JSON
        #self.save_results({"extractions": all_results}, str(Path(output_dir) / "all_extractions.json"))

        # ✅ Convert all to DataFrame
        df = pd.DataFrame(all_results)

        return df
    

    def list_excel_files(self,source_dir: str, extensions=("*.xlsx", "*.xlsm")) -> list[str]:
        """Return list of Excel files from a directory"""
        p = Path(source_dir)
        files = []
        for ext in extensions:
            files.extend(p.rglob(ext))  # rglob → recursive search
        return [str(f) for f in files]
